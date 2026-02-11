import os
import warnings

# Suppress openpyxl "no default style" warning when reading/writing Excel (harmless)
warnings.filterwarnings(
    'ignore',
    message="Workbook contains no default style",
    module='openpyxl.styles.stylesheet'
)

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

from config import INPUT_DIR, PROCESS_DIR, OUTPUT_DIR, OUTPUT_PPTX, GARTNER_EXCEL, INPUT_CURRENT
from utils import (
    read_input_file,
    upload_summary,
    map_cmdb,
    map_gartner,
)

app = Flask(__name__)
CORS(app)

os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(PROCESS_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def clean_output_files():
    """Remove generated PPT and Gartner Excel so fresh files are created on next run."""
    for filepath in (OUTPUT_PPTX, GARTNER_EXCEL):
        try:
            if os.path.isfile(filepath):
                os.remove(filepath)
        except OSError:
            pass


def save_upload(f):
    """Save uploaded file as input/current.<ext> (any filename allowed)."""
    ext = os.path.splitext(f.filename or '')[-1].lower()
    if ext not in ('.xlsx', '.xls', '.csv'):
        ext = '.xlsx'
    # Remove any existing current.*
    for e in ('.xlsx', '.xls', '.csv'):
        path = os.path.join(INPUT_DIR, INPUT_CURRENT + e)
        if os.path.isfile(path):
            os.remove(path)
    path = os.path.join(INPUT_DIR, INPUT_CURRENT + ext)
    f.save(path)
    clean_output_files()
    return path


@app.route('/api/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'error': 'No file selected'}), 400
    ext = os.path.splitext(f.filename)[-1].lower()
    if ext not in ('.xlsx', '.xls', '.csv'):
        return jsonify({'error': 'Allowed formats: .xlsx, .xls, .csv'}), 400
    try:
        path = save_upload(f)
        df = read_input_file(path)
        summary = upload_summary(df)
        return jsonify({'summary': summary})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/map-cmdb', methods=['POST'])
def api_map_cmdb():
    try:
        summary = map_cmdb()
        # Generate PPT from output/output.csv
        ppt_error = None
        try:
            from ppt.generator import generate_presentation
            ppt_ok = generate_presentation()
            if not ppt_ok:
                ppt_error = 'PPT generation returned False (check generator logs)'
                print(f'[WARN] {ppt_error}')
        except Exception as e:
            ppt_error = str(e)
            print(f'[ERROR] PPT generation failed: {e}')
            import traceback; traceback.print_exc()
        return jsonify({'summary': summary})
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 404
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def write_gartner_excel(summary):
    """Write Gartner summary to output/gartner_export.xlsx (Category, Sub-Category, Current Portfolio Apps, Recommended Gartner Apps, Market Leaders)."""
    import pandas as pd
    rows = []
    for r in (summary or []):
        rows.append({
            'Category': r.get('l1', ''),
            'Sub-Category': r.get('l2', ''),
            'Current Portfolio Apps (Top 5 by spend)': ', '.join(r['top5AppNames']) if isinstance(r.get('top5AppNames'), list) else str(r.get('top5AppNames', '')),
            'Recommended Gartner Apps': r.get('recommendedGartnerApps', ''),
            'Market Leaders': r.get('marketLeaders', ''),
        })
    df = pd.DataFrame(
        rows,
        columns=[
            'Category',
            'Sub-Category',
            'Current Portfolio Apps (Top 5 by spend)',
            'Recommended Gartner Apps',
            'Market Leaders',
        ],
    )
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with pd.ExcelWriter(GARTNER_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Border, Side

            ws = writer.sheets['Sheet1']

            # Auto-adjust column widths
            for i, col_name in enumerate(df.columns, 1):
                max_len = max(
                    df[col_name].astype(str).str.len().max() if len(df) else 0,
                    len(str(col_name))
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

            # Wrap text and add borders to all data cells (including header)
            thin = Side(style='thin', color='000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = border
        except Exception:
            # Formatting is best-effort; file still gets written even if this fails
            pass


@app.route('/api/map-gartner', methods=['POST'])
def api_map_gartner():
    try:
        summary = map_gartner()
        write_gartner_excel(summary)
        return jsonify({'summary': summary})
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 404
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download/ppt', methods=['GET'])
def download_ppt():
    if not os.path.isfile(OUTPUT_PPTX):
        return jsonify({'error': 'PPT not generated yet. Run Map CMDB first.'}), 404
    return send_file(OUTPUT_PPTX, as_attachment=True, download_name='output.pptx')


@app.route('/api/download/excel', methods=['GET'])
def download_excel():
    if not os.path.isfile(GARTNER_EXCEL):
        return jsonify({'error': 'Excel not generated yet. Run Gartner Leaders Mapping first.'}), 404
    return send_file(GARTNER_EXCEL, as_attachment=True, download_name='gartner_export.xlsx')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
